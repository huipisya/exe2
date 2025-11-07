import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
from datetime import datetime

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Имя директории для хранения файлов пользователей
DATA_DIR = os.getenv('DATA_DIR', 'user_data')
os.makedirs(DATA_DIR, exist_ok=True)

# ID чата для бэкапа
BACKUP_CHAT_ID = os.getenv('BACKUP_CHAT_ID')

# Глобальная переменная для application
app = None

# --- ФУНКЦИИ РАБОТЫ С ИНДИВИДУАЛЬНЫМИ ФАЙЛАМИ ---
def get_user_excel_file(user_id: int) -> str:
    return os.path.join(DATA_DIR, f'user_{user_id}.xlsx')

def init_user_excel(user_id: int):
    excel_file = get_user_excel_file(user_id)
    
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Посты"
        
        # ОБНОВЛЕНО: добавлена колонка "Процитировали"
        headers = ['№', 'Ссылка', 'Статус', 'Процитировали', 'Дата добавления']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        alignment_center = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = thin_border
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40  # Новая колонка
        ws.column_dimensions['E'].width = 22
        
        wb.save(excel_file)
        logger.info(f"Создан новый Excel файл для пользователя {user_id}")

def send_backup_for_user(user_id: int):
    if BACKUP_CHAT_ID:
        excel_file = get_user_excel_file(user_id)
        if os.path.exists(excel_file):
            try:
                import asyncio
                async def _send():
                    try:
                        with open(excel_file, 'rb') as f:
                            await app.bot.send_document(
                                chat_id=BACKUP_CHAT_ID,
                                document=f,
                                filename=f'backup_user_{user_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                            )
                        logger.info(f"Бэкап для пользователя {user_id} отправлен успешно")
                    except Exception as e:
                        logger.error(f"Ошибка отправки бэкапа для пользователя {user_id}: {e}")

                asyncio.create_task(_send())
            except Exception as e:
                logger.error(f"Ошибка подготовки бэкапа для пользователя {user_id}: {e}")
        else:
            logger.warning(f"Файл для бэкапа пользователя {user_id} не найден.")

def optimize_row_height(ws, row_num):
    """Оптимизация высоты строки на основе содержимого"""
    cell = ws[f'D{row_num}']
    if cell.value:
        lines = str(cell.value).count('\n') + 1
        ws.row_dimensions[row_num].height = max(15, lines * 15)

def add_post_to_excel(user_id: int, link: str, status=None):
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        init_user_excel(user_id)

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    row = ws.max_row + 1
    number = row - 1
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Настройка переноса текста для колонки "Процитировали"
    wrap_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    ws[f'A{row}'] = number
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'].value = link
    ws[f'B{row}'].hyperlink = link
    ws[f'B{row}'].font = Font(color="0563C1", underline="single")
    ws[f'B{row}'].border = thin_border
    
    ws[f'C{row}'] = status if status else ""
    ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'C{row}'].border = thin_border
    
    # НОВАЯ КОЛОНКА: Процитировали
    ws[f'D{row}'] = ""
    ws[f'D{row}'].alignment = wrap_alignment
    ws[f'D{row}'].border = thin_border
    
    ws[f'E{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws[f'E{row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'E{row}'].border = thin_border
    
    wb.save(excel_file)
    send_backup_for_user(user_id)
    return number

def update_post_status(user_id: int, link: str, status: str):
    """Обновление статуса времени выхода поста"""
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        logger.warning(f"Попытка обновления статуса в несуществующем файле для пользователя {user_id}")
        return False

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    for row in range(2, ws.max_row + 1):
        if ws[f'B{row}'].value == link:
            ws[f'C{row}'] = status
            wb.save(excel_file)
            send_backup_for_user(user_id)
            return True
    return False

def add_citation_to_post(user_id: int, link: str, citation: str):
    """Добавление цитирования к посту"""
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        logger.warning(f"Попытка добавления цитирования в несуществующем файле для пользователя {user_id}")
        return False

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    wrap_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    for row in range(2, ws.max_row + 1):
        if ws[f'B{row}'].value == link:
            current_citations = ws[f'D{row}'].value or ""
            
            # Добавляем новое цитирование через запятую
            if current_citations:
                new_citations = current_citations + ", " + citation
            else:
                new_citations = citation
            
            ws[f'D{row}'].value = new_citations
            ws[f'D{row}'].alignment = wrap_alignment
            
            # Оптимизация высоты строки
            optimize_row_height(ws, row)
            
            wb.save(excel_file)
            send_backup_for_user(user_id)
            return True
    return False
def link_exists_in_excel(user_id: int, link: str) -> bool:
    """Проверка существования ссылки в базе"""
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        return False

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if ws[f'B{row}'].value == link:
            return True
    return False

def get_post_info(user_id: int, link: str) -> dict:
    """Получение информации о посте"""
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        return None

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if ws[f'B{row}'].value == link:
            return {
                'number': ws[f'A{row}'].value,
                'status': ws[f'C{row}'].value,
                'citations': ws[f'D{row}'].value or "Пока никто не процитировал",
                'date': ws[f'E{row}'].value
            }
    return None

# --- ФУНКЦИИ ДЛЯ СОЗДАНИЯ КНОПОК ---
def get_time_options_keyboard():
    keyboard = [
        [InlineKeyboardButton("Вышли первыми", callback_data='status_1')],
        [InlineKeyboardButton("Вышли в течение часа", callback_data='status_2')],
        [InlineKeyboardButton("Вышли в течение 2-3 часов", callback_data='status_3')],
        [InlineKeyboardButton("Вышли больше, чем через 3 часа", callback_data='status_4')]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_edit_options_keyboard():
    """Клавиатура выбора действия для существующего поста"""
    keyboard = [
        [InlineKeyboardButton("✏️ Изменить время выхода", callback_data='edit_status')],
        [InlineKeyboardButton("📢 Добавить цитирование", callback_data='add_citation')],
        [InlineKeyboardButton("↩️ Отмена", callback_data='cancel')]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_new_link_keyboard():
    keyboard = [
        [InlineKeyboardButton("Отправить новую ссылку", callback_data='new_link')]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_after_add_keyboard():
    keyboard = [
        [InlineKeyboardButton("Отправить новую ссылку", callback_data='new_link')],
        [InlineKeyboardButton("Отправить актуальную базу данных", callback_data='export_db')]
    ]
    return InlineKeyboardMarkup(keyboard)

def extract_telegram_link(text: str) -> str:
    """Извлечение ссылки на Telegram из текста"""
    pattern = r'https?://(?:t\.me|telegram\.me)/(?:[a-zA-Z0-9_]+)(?:/[0-9]+)?(?:/[a-zA-Z0-9_]+)?'
    match = re.search(pattern, text)
    if match:
        return match.group(0)
    return ""

# --- ОБРАБОТЧИКИ ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    init_user_excel(user_id)
    await update.message.reply_text(
        f"👋 Привет, {update.effective_user.first_name}! Я бот для сохранения постов.\n\n"
        "Просто *перешли* мне пост из Telegram или отправь ссылку.\n\n"
        "Команды:\n"
        "/export - выгрузить твою базу данных в Excel\n"
        "/stats - статистика *твоих* постов"
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    init_user_excel(user_id)

    # Проверяем, ждём ли мы ввод цитирования
    if context.user_data.get('waiting_for_citation'):
        link = context.user_data.get('current_link')
        citation_text = update.message.text.strip()
        
        if citation_text:
            success = add_citation_to_post(user_id, link, citation_text)
            if success:
                await update.message.reply_text(
                    f"✅ Цитирование добавлено!\n\n"
                    f"Канал: {citation_text}",
                    reply_markup=get_after_add_keyboard()
                )
            else:
                await update.message.reply_text("❌ Ошибка при добавлении цитирования.")
        else:
            await update.message.reply_text("❌ Название канала не может быть пустым.")
        
        # Очищаем контекст
        context.user_data.pop('waiting_for_citation', None)
        context.user_data.pop('current_link', None)
        return

    # Обычная обработка ссылки
    text = update.message.text or update.message.caption or ""
    if text:
        link = extract_telegram_link(text)
        if link:
            if link_exists_in_excel(user_id, link):
                # Пост уже существует - предлагаем редактирование
                post_info = get_post_info(user_id, link)
                context.user_data['current_link'] = link
                
                message = (
                    f"📌 Этот пост уже есть в базе данных!\n\n"
                    f"Пост #{post_info['number']}\n"
                    f"Статус: {post_info['status']}\n"
                    f"Процитировали: {post_info['citations']}\n\n"
                    f"Что хотите изменить?"
                )
                
                reply_markup = get_edit_options_keyboard()
                await update.message.reply_text(message, reply_markup=reply_markup)
            else:
                # Новый пост
                context.user_data['current_link'] = link
                reply_markup = get_time_options_keyboard()
                await update.message.reply_text(
                    f"📌 Пост получен!\n\nСсылка: {link}\n\nУкажи, когда он вышел по кнопкам ниже",
                    reply_markup=reply_markup
                )
        else:
            await update.message.reply_text(
                "❌ Я не нашёл действительную ссылку на пост в Telegram в твоём сообщении.\n\n"
                "Отправь ссылку, перешли пост с комментарием или отправь медиа с подписью содержащей ссылку."
            )
    else:
        await update.message.reply_text(
            "❌ Я не нашёл действительную ссылку на пост в Telegram в твоём сообщении.\n\n"
            "Отправь ссылку, перешли пост с комментарием или отправь медиа с подписью содержащей ссылку."
        )

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id

    # Экспорт базы данных
    if query.data == 'export_db':
        excel_file = get_user_excel_file(user_id)
        if os.path.exists(excel_file):
            await query.message.reply_document(
                document=open(excel_file, 'rb'),
                filename=f'my_posts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        else:
            await query.edit_message_text("❌ Твоя база данных пуста. Добавь хотя бы один пост.")
        return

    # Новая ссылка
    if query.data == 'new_link':
        context.user_data.clear()
        await query.edit_message_text("✅ Готов принять новую ссылку. Отправь её сюда.")
        return

    # Отмена
    if query.data == 'cancel':
        context.user_data.clear()
        await query.edit_message_text("↩️ Действие отменено. Отправь новую ссылку.")
        return

    # Редактирование статуса
    if query.data == 'edit_status':
        link = context.user_data.get('current_link')
        if link:
            context.user_data['editing_status'] = True
            reply_markup = get_time_options_keyboard()
            await query.edit_message_text(
                f"✏️ Выбери новый статус времени выхода:",
                reply_markup=reply_markup
            )
        return

    # Добавление цитирования
    if query.data == 'add_citation':
        link = context.user_data.get('current_link')
        if link:
            context.user_data['waiting_for_citation'] = True
            await query.edit_message_text(
                "📢 Напиши название канала, который процитировал пост:"
            )
        return

    # Обработка выбора статуса
    link = context.user_data.get('current_link')
    if not link:
        await query.edit_message_text("❌ Ошибка: ссылка не найдена. Отправь ссылку заново.")
        return

    status_mapping = {
        'status_1': "Вышли первыми",
        'status_2': "Вышли в течение часа",
        'status_3': "Вышли в течение 2-3 часов",
        'status_4': "Вышли больше, чем через 3 часа"
    }

    selected_status = status_mapping.get(query.data)
    if selected_status:
        try:
            # Проверяем, это редактирование или новый пост
            if context.user_data.get('editing_status'):
                success = update_post_status(user_id, link, selected_status)
                if success:
                    await query.message.reply_text(
                        f"✅ Статус обновлён!\n\n"
                        f"Новый статус: {selected_status}",
                        reply_markup=get_after_add_keyboard()
                    )
                else:
                    await query.edit_message_text("❌ Ошибка при обновлении статуса.")
                context.user_data.pop('editing_status', None)
            else:
                # Добавление нового поста
                number = add_post_to_excel(user_id, link, selected_status)
                await query.message.reply_text(
                    f"✅ Пост #{number} добавлен в твою базу данных!\n\n"
                    f"Ссылка: {link}\n"
                    f"Статус: {selected_status}",
                    reply_markup=get_after_add_keyboard()
                )
            
            context.user_data.pop('current_link', None)
        except Exception as e:
            logger.error(f"Ошибка при работе с постом для пользователя {user_id}: {e}")
            await query.edit_message_text("❌ Произошла ошибка. Попробуй ещё раз.")
            context.user_data.clear()
    else:
        await query.edit_message_text("❌ Неизвестная команда. Попробуй снова.")
        context.user_data.clear()

async def export_database(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    excel_file = get_user_excel_file(user_id)
    if os.path.exists(excel_file):
        await update.message.reply_document(
            document=open(excel_file, 'rb'),
            filename=f'my_posts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    else:
        await update.message.reply_text("❌ Твоя база данных пуста. Добавь хотя бы один пост.")

async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        await update.message.reply_text("📊 Твоя база данных пуста.")
        return
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    total = ws.max_row - 1
    
    statuses = {}
    citations_count = 0
    
    for row in range(2, ws.max_row + 1):
        status = ws[f'C{row}'].value
        if status:
            statuses[status] = statuses.get(status, 0) + 1
        
        citations = ws[f'D{row}'].value
        if citations:
            citations_count += citations.count('\n') + 1
    
    message = f"📊 Статистика *твоих* постов:\n\n"
    message += f"Всего постов: {total}\n"
    message += f"Процитировано раз: {citations_count}\n\n"
    
    if statuses:
        message += "По статусам:\n"
        for status, count in statuses.items():
            message += f"• {status}: {count}\n"
    
    await update.message.reply_text(message)

def main():
    global app

    TOKEN = os.getenv("BOT_TOKEN")
    
    if not TOKEN:
        logger.error("Требуется переменная окружения BOT_TOKEN")
        return

    app = Application.builder().token(TOKEN).job_queue(None).build()
    
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("export", export_database))
    app.add_handler(CommandHandler("stats", stats))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(~filters.COMMAND, handle_message))
    
    logger.info("Бот запущен с функцией редактирования и отслеживания  цитирований!")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()